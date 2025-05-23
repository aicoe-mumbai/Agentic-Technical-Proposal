from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from typing import Dict, List, Optional
import os
import openpyxl

from Backend.app.models.models import (
    ProjectTemplate, TemplateRequest, TemplateResponse, AllTemplatesResponse,
    TemplateUpdateRequest
)
from Backend.app.db.database import save_template, get_all_templates, get_template_by_name, delete_template
from Backend.app.core.config import PROJECTS_DIR, BASE_DIR

router = APIRouter(prefix="/templates", tags=["templates"])

@router.get("/", response_model=AllTemplatesResponse)
async def list_templates():
    """List all available project templates"""
    templates = get_all_templates()
    return {"templates": templates}

@router.get("/{project_name}", response_model=TemplateResponse)
async def get_template(project_name: str):
    """Get a specific project template by name"""
    template = get_template_by_name(project_name)
    if not template:
        raise HTTPException(status_code=404, detail=f"Template '{project_name}' not found")
    return template

@router.post("/", response_model=TemplateResponse)
async def create_template(
    project_name: str = Form(...),
    project_TOC: str = Form(...),
    excel_file: Optional[UploadFile] = File(None)
):
    """Create a new project template with optional Excel file"""
    # Create project directory if it doesn't exist
    project_folder = os.path.join(PROJECTS_DIR, project_name)
    file_path = None
    
    if not os.path.exists(project_folder):
        os.makedirs(project_folder)
    
    # Save Excel file if provided
    if excel_file:
        file_path = os.path.join(project_folder, "template.xlsx")
        with open(file_path, "wb") as f:
            content = await excel_file.read()
            f.write(content)
    
    # Save template to database
    save_template(project_name, project_TOC, file_path or "")
    
    return {
        "project_name": project_name,
        "project_TOC": project_TOC,
        "file_path": file_path
    }

@router.delete("/{project_name}")
async def remove_template(project_name: str):
    """Delete a project template"""
    success = delete_template(project_name)
    if not success:
        raise HTTPException(status_code=404, detail=f"Template '{project_name}' not found")
    return {"message": f"Template '{project_name}' deleted successfully"}

@router.put("/{project_name}", response_model=TemplateResponse)
async def update_template_endpoint(
    project_name: str,
    project_TOC: str = Form(...),
    excel_file: Optional[UploadFile] = File(None)
):
    """Update an existing project template's TOC and optionally its Excel file."""
    existing_template = get_template_by_name(project_name)
    if not existing_template:
        raise HTTPException(status_code=404, detail=f"Template '{project_name}' not found")

    project_folder = os.path.join(PROJECTS_DIR, project_name)
    file_path_to_save = existing_template.get("file_path")

    if excel_file:
        if not os.path.exists(project_folder):
            os.makedirs(project_folder)
        
        new_file_path = os.path.join(project_folder, "template.xlsx")
        with open(new_file_path, "wb") as f:
            content = await excel_file.read()
            f.write(content)
        file_path_to_save = new_file_path
    
    save_template(project_name, project_TOC, file_path_to_save or "")
    
    return {
        "project_name": project_name,
        "project_TOC": project_TOC,
        "file_path": file_path_to_save
    }

@router.get("/{project_name}/data")
async def get_template_data(project_name: str):
    """Get Excel data from a template as key-value pairs"""
    print(f"[DEBUG] get_template_data called for: {project_name}")
    template = get_template_by_name(project_name)
    if not template:
        print(f"[DEBUG] Template '{project_name}' not found in DB.")
        raise HTTPException(status_code=404, detail=f"Template '{project_name}' not found")
    
    file_path_from_db = template.get("file_path")
    print(f"[DEBUG] BASE_DIR: {BASE_DIR}")
    print(f"[DEBUG] PROJECTS_DIR: {PROJECTS_DIR}")
    print(f"[DEBUG] file_path from DB: {file_path_from_db}")
    
    if not file_path_from_db:
        print(f"[DEBUG] No file_path defined for template '{project_name}'.")
        raise HTTPException(status_code=404, detail=f"Excel file for template '{project_name}' not found (no path defined)")

    absolute_file_path_to_check = os.path.join(BASE_DIR, file_path_from_db)
    print(f"[DEBUG] Checking absolute path: {absolute_file_path_to_check}")
    exists = os.path.exists(absolute_file_path_to_check)
    print(f"[DEBUG] os.path.exists({absolute_file_path_to_check}) result: {exists}")

    if not exists:
        path_relative_to_projects_dir = os.path.join(PROJECTS_DIR, os.path.basename(os.path.dirname(file_path_from_db)), os.path.basename(file_path_from_db))
        if os.path.isabs(file_path_from_db) and os.path.exists(file_path_from_db):
            print(f"[DEBUG] file_path_from_db is an absolute path and exists: {file_path_from_db}")
            absolute_file_path_to_check = file_path_from_db
            exists = True
        else:
            print(f"[DEBUG] File not found at {absolute_file_path_to_check}. Raising 404.")
            raise HTTPException(status_code=404, detail=f"Excel file for template '{project_name}' not found at expected path: {file_path_from_db}")
    
    try:
        wb = openpyxl.load_workbook(absolute_file_path_to_check)
        ws = wb['Sheet1']
        data_dict = {}
        for i in range(2, ws.max_row + 1):
            key = ws.cell(i, 1).value
            value = ws.cell(i, 2).value
            if key is not None and value is not None:
                data_dict[key] = value
        return data_dict
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}") 